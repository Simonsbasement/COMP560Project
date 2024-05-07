# Helper functions

import os
import numpy as np
import random
from colorama import Fore, Style
from openpyxl import Workbook, load_workbook

# Provide check for legal moves and top avalible row. 
# Input: int[][] b = board
# Return: [l, t]
#         bool[] l = boolean array indicating if the corresponding
#                    column is accepting legal move. 
#         int[] t  = int array indicating the top avalible row for
#                    each column. -1 if that column is full (illegal). 
def get_avalible_column(b):
    [n, m] = np.shape(np.array(b))
    l = []
    t = []
    for c in range(0, n):
        l.append(False)
        t.append(-1)
        for r in range(0, m):
            if b[c][r] == 0:
                l[-1] = True
                t[-1] = r
                break
    return np.array(l, dtype=bool), np.array(t, dtype=int)

# Print the board
# Input: int[][] b = board
def print_board(b):
    [n, m] = np.shape(np.array(b))
    for r in range(m-1, -1, -1):
        for c in range(0, n):
            match (b[c][r]):
                case 0:
                    print(f'{Fore.WHITE}_ {Style.RESET_ALL}', end='')
                case 1:
                    print(f'{Fore.RED}1 {Style.RESET_ALL}', end='')
                case 2:
                    print(f'{Fore.GREEN}2 {Style.RESET_ALL}', end='')
        print()
    return

# Play a move, return the new board; raise ValueError if trying to make an illegal move
# Input: int[][] b = board
#        int n = player to make the move
#        int c = which column to move
# Return int[][] b = new board
def make_move(b, n, c):
    b = np.copy(b)
    l, t = get_avalible_column(b)
    play = np.array(range(0, len(l)))[l]
    if c not in play:
        raise ValueError(f'Move {c} is an illegal move out of {play}.')
    else:
        b[c, t[c]] = n
        return b

# Backtrack move j and k
# Input: int[][] b = board
#        int j = one of the columns to backtrack
#        int k = the other column to backtrack
# Return int[][] b = new board
def backtrack(b, j, k):
    _, t = get_avalible_column(b)
    b[j][t[j]-1] = 0
    _, t = get_avalible_column(b)
    b[k][t[k]-1] = 0
    return b

# return 0 for no winner, 1/2 for respective winner
# evaled from 1) bot to top, then 2) left to right
# if there are more than 1 winner on board(idk how), return the 1st discovered winner
# Input: int[][] b = borad
#        int w = connect #
# Return: int winner = 1 or 2 if that player is winning
#                      0 if noone is currently winning
#                      3 if the board is full AND noone is winning -> a draw
def get_winner(b, w):
    [n, m] = np.shape(np.array(b))
    for c in range(0, n):
        for r in range(0, m):
            if b[c][r] == 0:
                continue
            # check in sequence: up->right->right-up->right-down
            conn = [1, 1, 1, 1]
            #up
            for rr in range(r+1, m):
                if (b[c][rr] == b[c][r]):
                    conn[0] = conn[0]+1
                else:
                    break
            #right
            for rr in range(c+1, n):
                if (b[rr][r] == b[c][r]):
                    conn[1] = conn[1]+1
                else:
                    break
            #right-up
            for rr in range(1, min(n-c, m-r)):
                if (b[c+rr][r+rr] == b[c][r]):
                    conn[2] = conn[2]+1
                else:
                    break
            #right-down
            for rr in range(1, min(n-c, r+1)):
                if (b[c+rr][r-rr] == b[c][r]):
                    conn[3] = conn[3]+1
                else:
                    break
            # if win, return winner
            for k in conn:
                if k == w:
                    return b[c][r]
                
    # potential draw -> check if full board
    l, t = get_avalible_column(b)
    if (np.any(l)):
        # no winner, not full board, return 0
        return 0
    else:
        # full board, no winner, draw -> return 3
        return 3

# Output the result of a game to an excel file. 
# Example output:
# agent1 = "Agent 1"
# agent2 = "Agent 2"
# winner = "Agent 1"
# match_time = "00:05:23"  # Example match time
# heuristic = "Heuristic 1"  # Example heuristic
# final_board = [[1, 0, 2, 0, 0, 0],
#                [2, 1, 1, 0, 0, 0],
#                [1, 2, 2, 0, 0, 0],
#                [2, 1, 2, 0, 0, 0],
#                [1, 2, 1, 0, 0, 0],
#                [2, 1, 1, 0, 0, 0]]  # Example final board state
def record_to_excel(agent1, agent2, winner, match_time, heuristic_1, heuristic_2, final_board, depth, tournament=False):
    if not tournament:
        file_name = "game_data.xlsx"
        headers = ["Agent 1", "Agent 2", "Winner", "Match Time", "Heuristic 1", "Heuristic 2", "Final Board", "Depth"]
    else:
        file_name = "game_data_tournament.xlsx"
        headers = ["Agent 1", "Agent 2", "Winrate", "Average Match Time", "Heuristic 1", "Heuristic 2", "Depth"]
        
    # Check if the file exists
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        # If not, create a new workbook if the file doesn't exist
        wb = Workbook()
        ws = wb.active
        # Define column headers for the new file
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

    if not tournament:
        # Convert final_board to a string
        final_board_str = '\n'.join([' '.join(map(str, row)) for row in final_board])
        # Append data to the next row
        row_data = [agent1, agent2, winner, match_time, heuristic_1, heuristic_2, final_board_str, depth]
        ws.append(row_data)
    else:
        # Append data to the next row
        row_data = [agent1, agent2, winner, match_time, heuristic_1, heuristic_2, depth]
        ws.append(row_data)
    
    # Save the workbook
    wb.save(file_name)
    if not tournament:
        print(f"Data appended to {file_name}")

# Helper for mcts: simulate a random game based on given board
def simulate_random_playout(b, n, w):
    current_player = n
    while True:
        result = get_winner(b, w)
        if result != 0:  # If game is over
            return result
        legal_moves = [col for col, is_legal in enumerate(get_avalible_column(b)[0]) if is_legal]
        if not legal_moves:  # No legal moves means a draw
            return 0
        move = random.choice(legal_moves)
        b = make_move(b, current_player, move)
        current_player = 3 - current_player  # Switch player
